# app/exporter/shopee_excel_exporter.py

import os
import shutil
from datetime import datetime
from typing import List
from openpyxl import load_workbook
from app.models.product import Product
from pathlib import Path
from urllib.parse import quote
from app.utils.image_tools import get_images_url
from app.utils.config import load_env
load_env()  # 確保讀到 .env 設定

BATCH_TEMPLATE_SOURCE = os.getenv("BATCH_MODULE_SOURCE")
BATCH_TEMPLATE_FILE = os.getenv("BATCH_MODULE_FILE")
CANDIDATES_ROOT = os.getenv("CANDIDATES_ROOT")

EXPORT_DIR = CANDIDATES_ROOT
EXPORT_SHEET_NAME = "上傳模板"



# === 主函式：將多筆產品轉換為 Shopee Excel ===
def export_products_to_excel(products: List[Product]) -> str:
    # 1. 複製模板
    template_path = os.path.join(BATCH_TEMPLATE_SOURCE, BATCH_TEMPLATE_FILE)
    ts = datetime.now().strftime("%y%m%d%H")
    export_filename = f"batch_upload_{ts}.xlsx"
    os.makedirs(EXPORT_DIR, exist_ok=True)
    export_path = os.path.join(EXPORT_DIR, export_filename)
    shutil.copy(template_path, export_path)

    # 2. 開啟 Excel 檔案
    wb = load_workbook(export_path)
    ws = wb[EXPORT_SHEET_NAME]
    current_row = 7  # 從第 7 列開始寫

    # 3. 處理每個 product → 展開規格組合
    for product in products:
        colors = product.colors or [None]
        sizes = product.sizes or [None]
        images = get_images_url(product.image_dir)

        main_image = images[0] if images else ""
        extra_images = images[1:9] if len(images) > 1 else []

        for color in colors:
            for size in sizes:
                variant_name = f"{color or ''}/{size or ''}".strip("/")

                # SKU & 規格識別碼自動生成
                product_sku = f"{product.stall_name}_{product.id}"
                variant_sku = f"{product.id}_{color or 'nospec'}_{size or 'nospec'}"
                variant_group_code = str(product.id)
                row = [
                    product.shopee_category_id or "",             # 分類
                    product.name or "",                           # 商品名稱
                    product.description or "Warning Empty",                    # 商品描述
                    product.min_purchase_qty or 1,                # 最低購買數量
                    product_sku,                                  # 主商品貨號
                    "", # variant_group_code,                           # 商品規格識別碼
                    "顏色" if color else "",                      # 規格名稱 1
                    color or "",                                  # 規格選項 1
                    main_image if (color or size) else "",                             # 規格圖片（若無主圖就空）
                    "尺寸" if size else "",                       # 規格名稱 2
                    size or "",                                   # 規格選項 2
                    int(product.price) if product.price else 9999,   # 價格（若空給預設 9999）
                    product.stock or 1,                           # 庫存（預設 1）
                    "",# variant_sku,                                  # 商品選項貨號
                    "", "", "",                                   # 尺寸表略
                    main_image or "",                             # 主商品圖片
                ]

                # 補圖片欄位（最多 8 張）
                for i in range(8):
                    row.append(extra_images[i] if i < len(extra_images) else "")

                # 後續欄位
                row += [
                    "", "", "", "",                              # 重量/尺寸略
                    "開啟" if "黑貓宅急便" in product.logistics_options else "",
                    "開啟" if "7-ELEVEN" in product.logistics_options else "",
                    "開啟" if "全家" in product.logistics_options else "",
                    "開啟" if "萊爾富" in product.logistics_options else "",
                    "開啟" if "蝦皮店到店" in product.logistics_options else "",
                    "開啟" if "店到家宅配" in product.logistics_options else "",
                    product.preparation_days or ""
                ]

                # 寫入一列
                for col_index, value in enumerate(row, start=1): 
                    ws.cell(row=current_row, column=col_index, value=value)

                current_row += 1  # 換下一列

    wb.save(export_path)
    return export_path
